from . column_info import Column


def __create_workbook__(columns,data):
    from  datetime import datetime
    from openpyxl.styles import Protection
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()

    ws = wb.active
    ws.title = "data"
    ws.protection.sheet = True
    if isinstance(columns, tuple):
        col_index = 1
        for col in columns:
            col_letter = get_column_letter(col_index)

            range = wb.create_named_range(col.field, ws, col_letter + ":" + col_letter)
            if isinstance(col, Column):
                ws.cell(1, col_index, col.caption)
            else:
                raise Exception("column {0} is not {1}".format(
                    col_index,
                    Column
                ))
            col_index = col_index + 1
        col_index = 1
        row_index = 2
        if isinstance(data, list):
            for row in data:
                col_index = 1
                for col in columns:
                    val = None

                    if isinstance(row, dict):
                        val = row.get(col.field,None)

                    else:
                        if hasattr(row,col.field):
                            val = getattr(row,col.field)
                    val_type = type(val)
                    if val_type in [str,unicode]:
                        ws.cell(row_index, col_index, val)

                    else:
                        ws.cell(row_index, col_index, val)
                    if col.format:
                        ws.cell(row_index, col_index).number_format = col.format
                    col_index = col_index + 1
                row_index = row_index + 1
    col_index = 1
    for col in columns:
        col_letter = get_column_letter(col_index)
        if col.format:
            ws.column_dimensions[col_letter].number_format = col.format
        ws.column_dimensions[col_letter].auto_size = True
        if not col.is_lock:
            ws.column_dimensions[col_letter].protection = Protection(
                locked=False
            )
        col_index = col_index + 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:" + get_column_letter(columns.__len__()) + "1"
    return wb


def save_to_stream(columns, data):
    from openpyxl.writer.excel import save_virtual_workbook
    wb = __create_workbook__(columns,data)
    ret = save_virtual_workbook(wb)
    return ret


def save_to_file(filename, columns, data):

    wb = __create_workbook__(columns, data)
    wb.save(filename)







